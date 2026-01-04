
#source .venv/bin/activate
#streamlit run app.py

import io
import os
import streamlit as st
import fitz  # PyMuPDF
import qrcode
from PIL import Image, ImageDraw, ImageFont
from openpyxl import load_workbook


# =========================
# App config
# =========================
st.set_page_config(page_title="Scantron QR 產生器", layout="wide")
st.title("📄 Scantron QR 產生器")


# =========================
# Font helpers (CJK)
# =========================
# macOS 常見中文字型（你目前環境是 macOS，所以 PingFang 基本上一定有）
CJK_FONT_PATHS = [
    "/System/Library/Fonts/PingFang.ttc",
    "/System/Library/Fonts/STHeiti Medium.ttc",
    "/Library/Fonts/Arial Unicode.ttf",
]

def find_cjk_font_path() -> str | None:
    for p in CJK_FONT_PATHS:
        if os.path.exists(p):
            return p
    return None

def get_cjk_pil_font(font_size_px: int) -> ImageFont.FreeTypeFont:
    """
    Pillow 用：顯示中文
    """
    p = find_cjk_font_path()
    if p:
        return ImageFont.truetype(p, font_size_px)
    return ImageFont.load_default()


# =========================
# QR + XLSX + PDF helpers
# =========================
def make_qr_pil(text: str, box_size: int = 8, border: int = 2) -> Image.Image:
    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=box_size,
        border=border,
    )
    qr.add_data(text)
    qr.make(fit=True)
    return qr.make_image(fill_color="black", back_color="white").convert("RGBA")


def read_people_from_xlsx(uploaded_xlsx) -> list[dict]:
    wb = load_workbook(uploaded_xlsx, data_only=True)
    ws = wb.active

    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header:
        return []

    header_map = {str(v).strip(): i for i, v in enumerate(header) if v is not None}

    if "student_id" not in header_map or "name" not in header_map:
        raise RuntimeError(
            f"XLSX 必須包含欄位：student_id, name（目前欄位：{list(header_map.keys())}）"
        )

    people = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        sid_val = row[header_map["student_id"]] if header_map["student_id"] < len(row) else None
        name_val = row[header_map["name"]] if header_map["name"] < len(row) else None

        sid = "" if sid_val is None else str(sid_val).strip()
        name = "" if name_val is None else str(name_val).strip()

        if not sid and not name:
            continue
        people.append({"id": sid, "name": name})

    return people


def render_pdf_first_page(pdf_bytes: bytes, scale: float):
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    page = doc[0]

    page_rect = page.rect  # ← 先存起來（還活著）

    pix = page.get_pixmap(matrix=fitz.Matrix(scale, scale))
    img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples).convert("RGBA")

    doc.close()
    return img, page_rect



def overlay_qr_on_preview(
    base_rgba: Image.Image,
    qr_text: str,
    x_pt: int,
    y_pt: int,
    qr_size_pt: int,
    scale: float,
    alpha: float,
    label_text: str,
    label_gap_pt: int = 5,
) -> Image.Image:
    """
    預覽用：在 PDF 第一頁 render 圖上疊「真的 QR」+ 文字（支援中文）
    """
    out = base_rgba.copy()

    # QR
    qr_img = make_qr_pil(qr_text)
    qr_px = max(1, int(qr_size_pt * scale))
    qr_img = qr_img.resize((qr_px, qr_px), resample=Image.NEAREST)

    # 半透明（只影響預覽）
    if alpha < 1.0:
        r, g, b, a = qr_img.split()
        a = a.point(lambda v: int(v * alpha))
        qr_img = Image.merge("RGBA", (r, g, b, a))

    x_px = int(x_pt * scale)
    y_px = int(y_pt * scale)

    out.alpha_composite(qr_img, (x_px, y_px))

    # 文字（中文）
    draw = ImageDraw.Draw(out)
    font = get_cjk_pil_font(font_size_px=max(14, int(16 * scale)))

    label_y_px = y_px + qr_px + int(label_gap_pt * scale)
    draw.text((x_px, label_y_px), label_text, fill=(0, 0, 0, 255), font=font)

    return out


def generate_output_pdf(
    template_bytes: bytes,
    people: list[dict],
    qr_size_pt: int,
    x_pt: int,
    y_pt: int,
    label_gap_pt: int,
    label_mode: str,  # "id" or "id_name"
) -> bytes:
    """
    輸出 PDF：每人一頁，插入 QR + 文字（支援中文）
    """
    cjk_font_path = find_cjk_font_path()
    if not cjk_font_path:
        raise RuntimeError(
            "找不到系統中文字型（macOS 建議 PingFang.ttc）。"
            "請修改 CJK_FONT_PATHS 指向一個可用的中文字型檔（.ttf/.ttc/.otf）。"
        )

    tpl = fitz.open(stream=template_bytes, filetype="pdf")
    base = tpl[0]
    w, h = base.rect.width, base.rect.height

    out = fitz.open()

    for p in people:
        sid, name = p["id"], p["name"]
        qr_text = f"{sid} {name}"

        page = out.new_page(width=w, height=h)
        page.show_pdf_page(page.rect, tpl, 0)

        # QR
        rect = fitz.Rect(x_pt, y_pt, x_pt + qr_size_pt, y_pt + qr_size_pt)
        qr_png_buf = io.BytesIO()
        make_qr_pil(qr_text).convert("RGB").save(qr_png_buf, format="PNG")
        page.insert_image(rect, stream=qr_png_buf.getvalue())

        # Label (中文)
        label_text = qr_text if label_mode == "id_name" else sid

        # 嵌入中文字型
        fontname = "cjkfont"
        page.insert_font(fontname=fontname, fontfile=cjk_font_path)

        # 用 textbox 比 insert_text 更穩（避免截斷）
        text_rect = fitz.Rect(
            x_pt,
            y_pt + qr_size_pt + label_gap_pt,
            x_pt + qr_size_pt * 3.0,              # 給寬一點，避免中文被截
            y_pt + qr_size_pt + label_gap_pt + 20
        )

        page.insert_textbox(
            text_rect,
            label_text,
            fontname=fontname,
            fontsize=10,
            color=(0, 0, 0),
            align=0,  # left
        )

    buf = io.BytesIO()
    out.save(buf)
    out.close()
    tpl.close()
    return buf.getvalue()


# =========================
# UI
# =========================
col_left, col_right = st.columns([1, 2])

with col_left:
    st.header("🔧 設定 / 上傳")

    pdf_file = st.file_uploader("上傳模板 PDF", type=["pdf"])
    xlsx_file = st.file_uploader("上傳名單 XLSX（欄位：student_id, name）", type=["xlsx"])

    st.divider()
    st.subheader("📌 位置與樣式")

    qr_size = st.slider("QR_SIZE (points)", 60, 260, 123, 1)
    x = st.slider("X (距左, points)", 0, 650, 389, 1)
    y = st.slider("Y (距上, points)", 0, 900, 125, 1)
    label_gap = st.slider("文字與 QR 距離 (points)", 0, 40, 0, 1)
    alpha = st.slider("QR 透明度（預覽用）", 0.2, 1.0, 0.75, 0.05)

    label_mode_ui = st.radio("QR 下方文字", ["只顯示學號", "學號 + 姓名"], index=0)
    label_mode = "id_name" if label_mode_ui == "學號 + 姓名" else "id"

    st.divider()
    st.subheader("⬇️ 輸出")
    out_name = st.text_input("輸出檔名", value="scantron_with_qr.pdf")

with col_right:
    st.header("👀 即時預覽（第一頁）")

    if not pdf_file:
        st.info("請先上傳模板 PDF")
    else:
        pdf_bytes = pdf_file.getvalue()

        preview_scale = 1.2
        base_img, _page_rect = render_pdf_first_page(pdf_bytes, scale=preview_scale)

        # 取名單第一筆當預覽內容（若沒有名單則用示例）
        sid_preview = "33"
        name_preview = "王小明"
        if xlsx_file:
            try:
                people_preview = read_people_from_xlsx(xlsx_file)
                if people_preview:
                    sid_preview = people_preview[0]["id"] or sid_preview
                    name_preview = people_preview[0]["name"] or name_preview
            except Exception as e:
                st.warning(f"名單讀取失敗（仍可預覽）：{e}")

        qr_text_preview = f"{sid_preview} {name_preview}"
        label_text_preview = qr_text_preview if label_mode == "id_name" else sid_preview

        composed = overlay_qr_on_preview(
            base_rgba=base_img,
            qr_text=qr_text_preview,
            x_pt=x,
            y_pt=y,
            qr_size_pt=qr_size,
            scale=preview_scale,
            alpha=alpha,
            label_text=label_text_preview,
            label_gap_pt=label_gap,
        )

        st.image(composed, use_container_width=True)
        st.caption(
            f"預覽使用名單第一筆（或示例）：student_id={sid_preview}，name={name_preview}。"
            "（X/Y/QR_SIZE 皆以 PDF points 計算，輸出與預覽一致）"
        )


# =========================
# Generate & Download
# =========================
st.divider()

if not (pdf_file and xlsx_file):
    st.warning("請同時上傳模板 PDF 與名單 XLSX 才能產生輸出。")
else:
    if st.button("🚀 產生 PDF"):
        try:
            with st.spinner("產生中..."):
                people = read_people_from_xlsx(xlsx_file)
                if not people:
                    st.error("名單是空的，請確認 XLSX 有資料列。")
                else:
                    out_pdf_bytes = generate_output_pdf(
                        template_bytes=pdf_file.getvalue(),
                        people=people,
                        qr_size_pt=qr_size,
                        x_pt=x,
                        y_pt=y,
                        label_gap_pt=label_gap,
                        label_mode=label_mode,
                    )

            st.success(f"完成！共 {len(people)} 頁")
            st.download_button(
                label="⬇️ 下載 PDF",
                data=out_pdf_bytes,
                file_name=out_name if out_name.lower().endswith(".pdf") else f"{out_name}.pdf",
                mime="application/pdf",
            )
        except Exception as e:
            st.error(str(e))
